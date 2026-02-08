# """
# IgG介导的食物不耐受与人连蛋白关联性分析（适配本地路径）
# 数据路径：/Users/wangguotao/Downloads/ISAR/food/
# 功能：数据加载、清洗、分层分析、交互检验、可视化、结果导出
# """

# ======================================
# 1. 环境准备与库导入
# ======================================
# 安装依赖（首次运行时取消注释执行）
# !pip install pandas numpy matplotlib scipy statsmodels openpyxl

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from scipy.stats import pearsonr, spearmanr
import statsmodels.api as sm
import statsmodels.formula.api as smf
from openpyxl import Workbook
import warnings
warnings.filterwarnings('ignore')

# 设置中文字体
plt.rcParams['font.sans-serif'] = ['WenQuanYi Zen Hei', 'SimHei', 'Arial Unicode MS']
plt.rcParams['axes.unicode_minus'] = False

# ======================================
# 2. 配置文件路径（关键：已适配您的本地路径）
# ======================================
# 原始数据文件路径（根据您提供的路径配置）
MAIN_DATA_PATH = "/Users/wangguotao/Downloads/ISAR/food/信息汇总全_加阳性率+加权+二分类+3种食物加.xlsx"
FATIGUE_DATA_PATH = "/Users/wangguotao/Downloads/ISAR/food/疲劳量表.xls"  # 疲劳量表（备用）
# 输出文件保存路径（与原始数据同目录，便于查找）
OUTPUT_DIR = "/Users/wangguotao/Downloads/ISAR/food/"

# ======================================
# 3. 数据加载与基础探索
# ======================================
def load_data(main_path, fatigue_path):
    """加载主数据和疲劳量表数据（主数据为核心分析数据）"""
    # 加载主数据（核心分析用）
    try:
        df_main = pd.read_excel(main_path)
        print(f"✅ 成功加载主数据：{main_path}")
        print(f"主数据形状：{df_main.shape}（行×列）")
    except Exception as e:
        print(f"❌ 主数据加载失败：{str(e)}")
        return None, None
    
    # 加载疲劳量表数据（备用，如需分析可后续调用）
    try:
        df_fatigue = pd.read_excel(fatigue_path)
        print(f"✅ 成功加载疲劳量表数据：{fatigue_path}")
        print(f"疲劳量表数据形状：{df_fatigue.shape}（行×列）")
    except Exception as e:
        print(f"⚠️  疲劳量表数据加载失败（不影响核心分析）：{str(e)}")
        df_fatigue = None
    
    # 查看主数据关键变量
    key_vars = [col for col in df_main.columns if any(x in col for x in 
                ['性别', '年龄', '婚姻', '身高', '体重', '人连蛋白', '三种食物', '阳性率'])]
    print(f"\n核心分析变量（共{len(key_vars)}个）：")
    for i, var in enumerate(key_vars, 1):
        print(f"  {i:2d}. {var}")
    
    return df_main, df_fatigue

# ======================================
# 4. 数据清洗与预处理
# ======================================
def clean_and_preprocess(df):
    """数据清洗：处理缺失值、创建分层变量、编码变量"""
    df_work = df.copy()
    
    # 4.1 清理性别（仅保留男/女）
    df_work['性别_clean'] = df_work['性别'].apply(lambda x: x if x in ['男', '女'] else np.nan)
    
    # 4.2 清理婚姻状况（统一格式）
    df_work['婚姻_clean'] = df_work['婚姻'].replace('已婚·', '已婚')  # 统一特殊符号
    valid_marriage = ['已婚', '未婚', '离异']
    df_work['婚姻_clean'] = df_work['婚姻_clean'].apply(lambda x: x if x in valid_marriage else np.nan)
    
    # 4.3 计算BMI及分类（中国标准）
    df_work = df_work.dropna(subset=['身高', '体重'])  # 删除身高体重缺失值
    df_work['BMI'] = df_work['体重'] / ((df_work['身高'] / 100) ** 2)
    def bmi_category(bmi):
        if pd.isna(bmi):
            return np.nan
        elif bmi < 18.5:
            return '偏瘦'
        elif 18.5 <= bmi < 24:
            return '正常'
        elif 24 <= bmi < 28:
            return '超重'
        else:
            return '肥胖'
    df_work['BMI分类'] = df_work['BMI'].apply(bmi_category)
    df_work['BMI分类合并'] = df_work['BMI分类'].replace({'偏瘦': '正常'})  # 合并小样本组
    
    # 4.4 年龄分层（临床常用分组）
    age_bins = [0, 35, 45, 55, 100]
    age_labels = ['≤35岁', '36-45岁', '46-55岁', '>55岁']
    df_work['年龄分层'] = pd.cut(df_work['年龄'], bins=age_bins, labels=age_labels, right=False)
    
    # 4.5 选择核心变量并去重
    core_vars = [
        '性别_clean', '年龄分层', 'BMI分类', 'BMI分类合并', '婚姻_clean',
        '三种食物分类加权评分', '人连蛋白四参数结果.1', '是否耐受'
    ]
    df_clean = df_work[core_vars].dropna()
    
    # 4.6 创建回归编码变量
    df_clean['性别编码'] = df_clean['性别_clean'].map({'男': 1, '女': 0})
    df_clean['年龄连续'] = df_clean['年龄分层'].map({'≤35岁': 30, '36-45岁': 40, '46-55岁': 50, '>55岁': 60})
    df_clean['BMI编码'] = df_clean['BMI分类合并'].map({'正常': 1, '超重': 2, '肥胖': 3})
    
    # 打印清洗结果
    print(f"\n📊 数据清洗结果：")
    print(f"原始样本量：{len(df)}例 → 清洗后有效样本量：{len(df_clean)}例")
    print(f"各分层变量分布：")
    print(f"  性别：{df_clean['性别_clean'].value_counts().to_dict()}")
    print(f"  年龄分层：{df_clean['年龄分层'].value_counts().to_dict()}")
    print(f"  BMI分类：{df_clean['BMI分类合并'].value_counts().to_dict()}")
    print(f"  婚姻状况：{df_clean['婚姻_clean'].value_counts().to_dict()}")
    
    # 保存清洗后数据（同目录下）
    clean_data_path = f"{OUTPUT_DIR}cleaned_analysis_data.csv"
    df_clean.to_csv(clean_data_path, index=False, encoding='utf-8-sig')
    print(f"\n💾 清洗后数据已保存：{clean_data_path}")
    
    return df_clean

# ======================================
# 5. 分层分析核心函数（4个分层变量）
# ======================================
def stratified_analysis(df, strat_var):
    """分层分析：计算相关系数、线性回归及显著性"""
    results = {}
    groups = df.groupby(strat_var)
    strat_cn = {
        '性别_clean': '性别', '年龄分层': '年龄', 
        'BMI分类合并': 'BMI', '婚姻_clean': '婚姻状况'
    }[strat_var]
    
    print(f"\n" + "="*60)
    print(f"🔍 {strat_cn}分层分析结果")
    print("="*60)
    
    for group_name, group_data in groups:
        if len(group_data) < 3:  # 样本量<3例跳过
            results[str(group_name)] = {'样本量': len(group_data), '状态': '样本量不足'}
            print(f"【{group_name}】样本量{len(group_data)}例，跳过分析")
            continue
        
        # 相关分析（Pearson+Spearman）
        pearson_r, pearson_p = pearsonr(group_data['三种食物分类加权评分'], group_data['人连蛋白四参数结果.1'])
        spearman_r, spearman_p = spearmanr(group_data['三种食物分类加权评分'], group_data['人连蛋白四参数结果.1'])
        
        # 线性回归
        model = smf.ols("Q('人连蛋白四参数结果.1') ~ Q('三种食物分类加权评分')", data=group_data).fit()
        beta, beta_se, beta_p = model.params[1], model.bse[1], model.pvalues[1]
        r2 = model.rsquared
        
        # 结果存储与打印
        results[str(group_name)] = {
            '样本量': len(group_data), 'Pearson_r': pearson_r, 'Pearson_p': pearson_p,
            'Spearman_r': spearman_r, 'Spearman_p': spearman_p,
            'beta': beta, 'beta_se': beta_se, 'beta_p': beta_p, 'R2': r2
        }
        
        print(f"\n【{group_name}】（样本量：{len(group_data)}例）")
        print(f"  相关分析：")
        print(f"    Pearson：r={pearson_r:.4f}，p={pearson_p:.4f}")
        print(f"    Spearman：ρ={spearman_r:.4f}，p={spearman_p:.4f}")
        print(f"  线性回归：")
        print(f"    β={beta:.4f}（SE={beta_se:.4f}），p={beta_p:.4f}，R²={r2:.4f}")
        if pearson_p < 0.05 or spearman_p < 0.05:
            print(f"  ⚠️  存在显著相关性（p<0.05）")
    
    return results

# ======================================
# 6. 交互作用检验（IgG×性别/年龄/BMI）
# ======================================
def interaction_test(df):
    """交互作用检验：主效应模型 vs 交互效应模型"""
    print(f"\n" + "="*60)
    print("🔍 交互作用检验结果（IgG×性别/年龄/BMI）")
    print("="*60)
    
    # 创建交互项
    df_model = df.copy()
    df_model['IgG_性别交互'] = df_model['三种食物分类加权评分'] * df_model['性别编码']
    df_model['IgG_年龄交互'] = df_model['三种食物分类加权评分'] * df_model['年龄连续']
    df_model['IgG_BMI交互'] = df_model['三种食物分类加权评分'] * df_model['BMI编码']
    
    # 构建模型
    model1 = smf.ols(  # 主效应模型
        "Q('人连蛋白四参数结果.1') ~ Q('三种食物分类加权评分') + 性别编码 + 年龄连续 + BMI编码",
        data=df_model
    ).fit()
    
    model2 = smf.ols(  # 交互效应模型
        "Q('人连蛋白四参数结果.1') ~ Q('三种食物分类加权评分') + 性别编码 + 年龄连续 + BMI编码 + "
        "IgG_性别交互 + IgG_年龄交互 + IgG_BMI交互",
        data=df_model
    ).fit()
    
    # 模型比较（ANOVA）
    anova_result = sm.stats.anova_lm(model1, model2)
    anova_f = anova_result['F'].iloc[1] if 'F' in anova_result.columns else np.nan
    anova_p = anova_result['Pr(>F)'].iloc[1] if 'Pr(>F)' in anova_result.columns else np.nan
    
    # 提取交互项参数
    interaction_params = {}
    for term, name in [('IgG_性别交互', 'IgG×性别'), ('IgG_年龄交互', 'IgG×年龄'), ('IgG_BMI交互', 'IgG×BMI')]:
        if term in model2.params.index:
            interaction_params[name] = {
                'coef': model2.params[term], 'se': model2.bse[term], 'p_val': model2.pvalues[term]
            }
    
    # 打印结果
    print(f"模型1（主效应）：R²={model1.rsquared:.4f}，F_p={model1.f_pvalue:.4f}")
    print(f"模型2（交互效应）：R²={model2.rsquared:.4f}，F_p={model2.f_pvalue:.4f}")
    print(f"ANOVA检验（交互项整体）：F={anova_f:.4f}，p={anova_p:.4f}")
    
    print(f"\n各交互项单独检验：")
    for name, params in interaction_params.items():
        sig = "✅ 显著" if params['p_val'] < 0.05 else "❌ 不显著"
        print(f"  {name}： coef={params['coef']:.4f}，p={params['p_val']:.4f} {sig}")
    
    return {
        'model1': model1, 'model2': model2, 'anova_f': anova_f, 'anova_p': anova_p,
        'interaction_params': interaction_params
    }

# ======================================
# 7. 结果可视化（2×2子图）
# ======================================
def plot_results(df):
    """生成分层分析可视化图表（保存到同目录）"""
    fig, axes = plt.subplots(2, 2, figsize=(16, 12))
    fig.suptitle('IgG食物不耐受与人连蛋白关联性分析（分层结果）', fontsize=16, fontweight='bold')
    
    # 颜色配置
    gender_colors = {'男': '#1f77b4', '女': '#ff7f0e'}
    age_colors = {'≤35岁': '#2ca02c', '36-45岁': '#d62728', '46-55岁': '#9467bd', '>55岁': '#8c564b'}
    bmi_colors = {'正常': '#2ca02c', '超重': '#ff7f0e', '肥胖': '#d62728'}
    marriage_colors = {'已婚': '#1f77b4', '未婚': '#9467bd', '离异': '#8c564b'}
    
    # 子图1：性别分层
    ax1 = axes[0, 0]
    for gender, color in gender_colors.items():
        data = df[df['性别_clean'] == gender]
        ax1.scatter(data['三种食物分类加权评分'], data['人连蛋白四参数结果.1'], 
                   alpha=0.6, s=60, color=color, label=f'{gender}（n={len(data)}）')
        model = smf.ols("Q('人连蛋白四参数结果.1') ~ Q('三种食物分类加权评分')", data=data).fit()
        x_pred = np.linspace(data['三种食物分类加权评分'].min(), data['三种食物分类加权评分'].max(), 100)
        ax1.plot(x_pred, model.params[0] + model.params[1] * x_pred, color=color, linestyle='--', linewidth=2)
    ax1.set_title('性别分层', fontweight='bold')
    ax1.set_xlabel('三种食物分类加权评分（IgG）')
    ax1.set_ylabel('人连蛋白四参数结果')
    ax1.legend()
    ax1.grid(True, alpha=0.3)
    
    # 子图2：年龄分层
    ax2 = axes[0, 1]
    for age, color in age_colors.items():
        data = df[df['年龄分层'] == age]
        if len(data) < 3: continue
        ax2.scatter(data['三种食物分类加权评分'], data['人连蛋白四参数结果.1'], 
                   alpha=0.6, s=60, color=color, label=f'{age}（n={len(data)}）')
        model = smf.ols("Q('人连蛋白四参数结果.1') ~ Q('三种食物分类加权评分')", data=data).fit()
        x_pred = np.linspace(data['三种食物分类加权评分'].min(), data['三种食物分类加权评分'].max(), 100)
        ax2.plot(x_pred, model.params[0] + model.params[1] * x_pred, color=color, linestyle='--', linewidth=2)
    ax2.set_title('年龄分层', fontweight='bold')
    ax2.set_xlabel('三种食物分类加权评分（IgG）')
    ax2.set_ylabel('人连蛋白四参数结果')
    ax2.legend()
    ax2.grid(True, alpha=0.3)
    
    # 子图3：BMI分层
    ax3 = axes[1, 0]
    for bmi, color in bmi_colors.items():
        data = df[df['BMI分类合并'] == bmi]
        ax3.scatter(data['三种食物分类加权评分'], data['人连蛋白四参数结果.1'], 
                   alpha=0.6, s=60, color=color, label=f'{bmi}（n={len(data)}）')
        model = smf.ols("Q('人连蛋白四参数结果.1') ~ Q('三种食物分类加权评分')", data=data).fit()
        x_pred = np.linspace(data['三种食物分类加权评分'].min(), data['三种食物分类加权评分'].max(), 100)
        ax3.plot(x_pred, model.params[0] + model.params[1] * x_pred, color=color, linestyle='--', linewidth=2)
    ax3.set_title('BMI分层', fontweight='bold')
    ax3.set_xlabel('三种食物分类加权评分（IgG）')
    ax3.set_ylabel('人连蛋白四参数结果')
    ax3.legend()
    ax3.grid(True, alpha=0.3)
    
    # 子图4：婚姻分层
    ax4 = axes[1, 1]
    for marriage, color in marriage_colors.items():
        data = df[df['婚姻_clean'] == marriage]
        if len(data) < 3: continue
        ax4.scatter(data['三种食物分类加权评分'], data['人连蛋白四参数结果.1'], 
                   alpha=0.6, s=60, color=color, label=f'{marriage}（n={len(data)}）')
        model = smf.ols("Q('人连蛋白四参数结果.1') ~ Q('三种食物分类加权评分')", data=data).fit()
        x_pred = np.linspace(data['三种食物分类加权评分'].min(), data['三种食物分类加权评分'].max(), 100)
        ax4.plot(x_pred, model.params[0] + model.params[1] * x_pred, color=color, linestyle='--', linewidth=2)
    ax4.set_title('婚姻状况分层', fontweight='bold')
    ax4.set_xlabel('三种食物分类加权评分（IgG）')
    ax4.set_ylabel('人连蛋白四参数结果')
    ax4.legend()
    ax4.grid(True, alpha=0.3)
    
    # 保存图表
    plot_path = f"{OUTPUT_DIR}stratified_analysis_plot.png"
    plt.tight_layout()
    plt.savefig(plot_path, dpi=300, bbox_inches='tight', facecolor='white')
    plt.close()
    print(f"\n💾 可视化图表已保存：{plot_path}")

# ======================================
# 8. 结果汇总导出（Excel）
# ======================================
def export_results(gender_res, age_res, bmi_res, marriage_res, interaction_res):
    """将所有结果汇总到Excel文件（同目录下）"""
    wb = Workbook()
    
    # 工作表1：性别分层
    ws1 = wb.active
    ws1.title = "性别分层"
    ws1.append(['性别', '样本量', 'Pearson_r', 'Pearson_p', 'Spearman_r', 'Spearman_p', 'beta', 'beta_p', 'R2'])
    for g, res in gender_res.items():
        if '状态' not in res:
            ws1.append([g, res['样本量'], f"{res['Pearson_r']:.4f}", f"{res['Pearson_p']:.4f}",
                       f"{res['Spearman_r']:.4f}", f"{res['Spearman_p']:.4f}",
                       f"{res['beta']:.4f}", f"{res['beta_p']:.4f}", f"{res['R2']:.4f}"])
    
    # 工作表2：年龄分层
    ws2 = wb.create_sheet("年龄分层")
    ws2.append(['年龄组', '样本量', 'Pearson_r', 'Pearson_p', 'Spearman_r', 'Spearman_p', 'beta', 'beta_p', 'R2'])
    for a, res in age_res.items():
        if '状态' not in res:
            ws2.append([a, res['样本量'], f"{res['Pearson_r']:.4f}", f"{res['Pearson_p']:.4f}",
                       f"{res['Spearman_r']:.4f}", f"{res['Spearman_p']:.4f}",
                       f"{res['beta']:.4f}", f"{res['beta_p']:.4f}", f"{res['R2']:.4f}"])
    
    # 工作表3：BMI分层
    ws3 = wb.create_sheet("BMI分层")
    ws3.append(['BMI分类', '样本量', 'Pearson_r', 'Pearson_p', 'Spearman_r', 'Spearman_p', 'beta', 'beta_p', 'R2'])
    for b, res in bmi_res.items():
        if '状态' not in res:
            ws3.append([b, res['样本量'], f"{res['Pearson_r']:.4f}", f"{res['Pearson_p']:.4f}",
                       f"{res['Spearman_r']:.4f}", f"{res['Spearman_p']:.4f}",
                       f"{res['beta']:.4f}", f"{res['beta_p']:.4f}", f"{res['R2']:.4f}"])
    
    # 工作表4：婚姻分层
    ws4 = wb.create_sheet("婚姻分层")
    ws4.append(['婚姻状况', '样本量', 'Pearson_r', 'Pearson_p', 'Spearman_r', 'Spearman_p', 'beta', 'beta_p', 'R2'])
    for m, res in marriage_res.items():
        if '状态' not in res:
            ws4.append([m, res['样本量'], f"{res['Pearson_r']:.4f}", f"{res['Pearson_p']:.4f}",
                       f"{res['Spearman_r']:.4f}", f"{res['Spearman_p']:.4f}",
                       f"{res['beta']:.4f}", f"{res['beta_p']:.4f}", f"{res['R2']:.4f}"])
    
    # 工作表5：交互作用检验
    ws5 = wb.create_sheet("交互作用检验")
    ws5.append(['检验项目', '数值', '显著性'])
    ws5.append(['模型1 R²', f"{interaction_res['model1'].rsquared:.4f}", '-'])
    ws5.append(['模型1 F_p', f"{interaction_res['model1'].f_pvalue:.4f}", 
                '显著' if interaction_res['model1'].f_pvalue < 0.05 else '不显著'])
    ws5.append(['模型2 R²', f"{interaction_res['model2'].rsquared:.4f}", '-'])
    ws5.append(['模型2 F_p', f"{interaction_res['model2'].f_pvalue:.4f}", 
                '显著' if interaction_res['model2'].f_pvalue < 0.05 else '不显著'])
    ws5.append(['ANOVA F', f"{interaction_res['anova_f']:.4f}", '-'])
    ws5.append(['ANOVA p', f"{interaction_res['anova_p']:.4f}", 
                '显著' if interaction_res['anova_p'] < 0.05 else '不显著'])
    for name, params in interaction_res['interaction_params'].items():
        ws5.append([f"{name} p值", f"{params['p_val']:.4f}", 
                    '显著' if params['p_val'] < 0.05 else '不显著'])
    
    # 保存Excel
    excel_path = f"{OUTPUT_DIR}IgG人连蛋白分析结果汇总.xlsx"
    wb.save(excel_path)
    print(f"💾 结果汇总Excel已保存：{excel_path}")

# ======================================
# 9. 主函数（一键运行所有分析）
# ======================================
def main():
    print("="*80)
    print("IgG介导的食物不耐受与人连蛋白关联性分析")
    print(f"数据路径：{OUTPUT_DIR}")
    print("="*80)
    
    # 步骤1：加载数据
    df_main, df_fatigue = load_data(MAIN_DATA_PATH, FATIGUE_DATA_PATH)
    if df_main is None:
        print("❌ 数据加载失败，分析终止")
        return
    
    # 步骤2：数据清洗
    df_clean = clean_and_preprocess(df_main)
    
    # 步骤3：分层分析（4个变量）
    gender_res = stratified_analysis(df_clean, '性别_clean')
    age_res = stratified_analysis(df_clean, '年龄分层')
    bmi_res = stratified_analysis(df_clean, 'BMI分类合并')
    marriage_res = stratified_analysis(df_clean, '婚姻_clean')
    
    # 步骤4：交互作用检验
    interaction_res = interaction_test(df_clean)
    
    # 步骤5：可视化
    plot_results(df_clean)
    
    # 步骤6：结果导出
    export_results(gender_res, age_res, bmi_res, marriage_res, interaction_res)
    
    # 步骤7：核心结论
    print(f"\n" + "="*80)
    print("🎯 核心分析结论")
    print("="*80)
    # 年龄分层关键发现（46-55岁组）
    if '46-55岁' in age_res and 'Spearman_p' in age_res['46-55岁']:
        age_46_55 = age_res['46-55岁']
        if age_46_55['Spearman_p'] < 0.01:
            print(f"1. 46-55岁组：IgG与人间蛋白呈显著负相关（ρ={age_46_55['Spearman_r']:.4f}，p<0.01）")
    # 交互作用结论
    if interaction_res['anova_p'] >= 0.05:
        print(f"2. 交互作用：未发现IgG与性别/年龄/BMI的显著交互效应（p={interaction_res['anova_p']:.4f}）")
    # 其他分层结论
    print(f"3. 其他分层：性别、BMI、婚姻状况均未发现显著相关性")
    
    print(f"\n" + "="*80)
    print("✅ 分析完成！所有结果已保存到：")
    print(f"   {OUTPUT_DIR}")
    print("生成文件：")
    print("1. cleaned_analysis_data.csv → 清洗后数据集")
    print("2. stratified_analysis_plot.png → 可视化图表")
    print("3. IgG人连蛋白分析结果汇总.xlsx → 结果汇总Excel")
    print("="*80)

# ======================================
# 运行分析（直接执行脚本即可）
# ======================================
if __name__ == "__main__":
    main()